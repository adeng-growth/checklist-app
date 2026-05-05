#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
中小企业增长自检清单 - Vercel Python Serverless 入口
只暴露一个 handler(request) 函数给 Vercel
"""

import os
import json
import base64
import hashlib
import hmac
import requests as req_lib
from datetime import datetime
from io import BytesIO

# ===== 配置 =====
AIRTABLE_API_KEY = os.getenv('AIRTABLE_API_KEY', '')
AIRTABLE_BASE_ID = os.getenv('AIRTABLE_BASE_ID', '')
AIRTABLE_TABLE_NAME = 'Submissions'
ADMIN_PASSWORD = os.getenv('ADMIN_PASSWORD', 'adeng2026')
SECRET_KEY = os.getenv('SECRET_KEY', 'adeng-checklist-2026')

_HERE = os.path.dirname(os.path.abspath(__file__))


def _airtable_headers():
    return {
        'Authorization': 'Bearer ' + AIRTABLE_API_KEY,
        'Content-Type': 'application/json'
    }


def _airtable_url():
    return 'https://api.airtable.com/v0/' + AIRTABLE_BASE_ID + '/' + AIRTABLE_TABLE_NAME


def _resp(status, headers, body):
    """构造 Vercel 响应 dict"""
    return {'statusCode': status, 'headers': headers, 'body': body}


def _json_resp(data, status=200):
    return _resp(status, {'Content-Type': 'application/json; charset=utf-8'}, json.dumps(data, ensure_ascii=False))


def _html_resp(body, status=200):
    return _resp(status, {'Content-Type': 'text/html; charset=utf-8'}, body)


def _redirect_resp(location):
    return _resp(302, {'Location': location}, '')


# ===== Session =====
SESSION_COOKIE = 'ad_sess'
_SESSION_KEY = SECRET_KEY.encode('utf-8')


def _parse_session(cookies_str):
    raw = {}
    for pair in (cookies_str or '').split(';'):
        pair = pair.strip()
        if '=' in pair:
            k, v = pair.split('=', 1)
            raw[k.strip()] = v.strip()
    val = raw.get(SESSION_COOKIE, '')
    if not val or '.' not in val:
        return {}
    parts = val.split('.', 1)
    if len(parts) != 2:
        return {}
    data_b64, sig = parts
    try:
        data_json = base64.b64decode(data_b64).decode('utf-8')
        expect = hmac.new(_SESSION_KEY, data_json.encode('utf-8'), hashlib.sha256).hexdigest()[:16]
        if sig == expect:
            return json.loads(data_json)
    except Exception:
        pass
    return {}


def _make_cookie(data):
    data_json = json.dumps(data, ensure_ascii=False)
    data_b64 = base64.b64encode(data_json.encode('utf-8')).decode('utf-8')
    sig = hmac.new(_SESSION_KEY, data_json.encode('utf-8'), hashlib.sha256).hexdigest()[:16]
    return SESSION_COOKIE + '=' + data_b64 + '.' + sig + '; Path=/; HttpOnly; Max-Age=86400'


def _is_auth(cookies_str):
    return _parse_session(cookies_str).get('auth') is True


# ===== 静态文件 =====
def _static(filename):
    root_dir = os.path.dirname(_HERE)
    filepath = os.path.join(root_dir, filename)
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            return f.read()
    except Exception:
        return None


# ===== HTML 模板 =====
LOGIN_HTML = '''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>后台登录</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:-apple-system,BlinkMacSystemFont,"PingFang SC","Microsoft YaHei",sans-serif;background:linear-gradient(135deg,#1a365d,#2d5a87);min-height:100vh;display:flex;align-items:center;justify-content:center}
.box{background:#fff;padding:40px 36px;border-radius:20px;box-shadow:0 20px 60px rgba(0,0,0,.3);width:92%;max-width:380px}
h2{font-size:22px;color:#1a365d;margin-bottom:8px;text-align:center}
.sub{font-size:13px;color:#999;margin-bottom:28px;text-align:center}
.err{background:#ffebee;color:#c62828;padding:10px 16px;border-radius:8px;margin-bottom:16px;font-size:13px;text-align:center}
input{width:100%;padding:12px 16px;border:1.5px solid #e5e5e5;border-radius:10px;font-size:15px;outline:none;transition:border .2s;margin-bottom:20px}
input:focus{border-color:#2d5a87}
button{width:100%;padding:14px;background:linear-gradient(135deg,#1a365d,#2d5a87);color:#fff;border:none;border-radius:10px;font-size:16px;font-weight:600;cursor:pointer;transition:transform .15s}
button:active{transform:scale(.97)}
.ft{font-size:11px;color:#ccc;text-align:center;margin-top:20px}
</style></head>
<body>
<div class="box">
  <h2>&#128274; 后台登录</h2>
  <p class="sub">中小企业增长自检清单</p>
  {{ERROR}}
  <form method="POST"><input type="password" name="password" placeholder="请输入后台密码" required autofocus>
  <button type="submit">登 录</button></form>
  <p class="ft">ADENG CHECKLIST 2026</p>
</div></body></html>'''

ADMIN_HTML = '''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>管理后台</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:-apple-system,BlinkMacSystemFont,"PingFang SC","Microsoft YaHei",sans-serif;background:#f5f7fa}
.top{background:linear-gradient(135deg,#1a365d,#2d5a87);color:#fff;padding:16px 24px;display:flex;justify-content:space-between;align-items:center}
.top h1{font-size:18px}
.top a{color:#fff;text-decoration:none;font-size:13px;background:rgba(255,255,255,.2);padding:6px 14px;border-radius:8px}
.container{max-width:1100px;margin:24px auto;padding:0 16px}
.card{background:#fff;border-radius:16px;padding:24px;box-shadow:0 2px 12px rgba(0,0,0,.08);margin-bottom:20px}
.filters{display:flex;gap:10px;margin-bottom:16px;flex-wrap:wrap}
.filters input,.filters select{padding:8px 12px;border:1.5px solid #e5e5e5;border-radius:8px;font-size:14px}
.filters button{padding:8px 18px;background:#1a365d;color:#fff;border:none;border-radius:8px;cursor:pointer}
table{width:100%;border-collapse:collapse;font-size:13px}
th{background:#f0f4f8;padding:10px 8px;text-align:center;font-weight:600;color:#1a365d;border-bottom:2px solid #e5e5e5}
td{padding:10px 8px;text-align:center;border-bottom:1px solid #f0f0f0}
tr:hover{background:#f8fafc}
.tag{padding:3px 10px;border-radius:20px;font-size:12px;font-weight:600}
.tag-danger{background:#ffebee;color:#c62828}
.tag-warning{background:#fff8e1;color:#e65100}
.tag-info{background:#e3f2fd;color:#1565c0}
.tag-success{background:#e8f5e9;color:#2e7d32}
.tag-primary{background:#f3e5f5;color:#6a1b9a}
</style></head>
<body>
<div class="top"><h1>&#128218; 提交记录</h1><div><a href="/api/export/excel">导出Excel</a><a href="/admin/logout" style="margin-left:8px">退出</a></div></div>
<div class="container">
<div class="card">
  <div class="filters">
    <input type="text" id="kw" placeholder="搜索公司/姓名">
    <select id="rtype"><option value="">全部类型</option><option>危机型</option><option>问题型</option><option>潜力型</option><option>健康型</option><option>优秀型</option></select>
    <button onclick="load()">搜索</button>
    <span style="margin-left:auto;line-height:35px;color:#666" id="total"></span>
  </div>
  <div style="overflow-x:auto">
  <table>
  <tr><th>姓名</th><th>公司</th><th>电话</th><th>行业</th><th>总分</th><th>商业模式</th><th>品牌定位</th><th>流量获客</th><th>转化成交</th><th>组织系统</th><th>类型</th><th>时间</th></tr>
  <tbody id="tbody"></tbody>
  </table>
  </div>
</div></div>
<script>
let cache=[];
function load(){
  fetch('/api/records').then(r=>r.json()).then(d=>{
    cache=d.records||[];
    render();
  });
}
function render(){
  let kw=document.getElementById('kw').value.trim().toLowerCase();
  let rt=document.getElementById('rtype').value;
  let f=cache.filter(r=>{
    if(kw && !(r.name||'').toLowerCase().includes(kw) && !(r.company||'').toLowerCase().includes(kw)) return false;
    if(rt && r.result_type!=rt) return false;
    return true;
  });
  document.getElementById('total').innerText='共 '+f.length+' 条';
  let html=f.map(r=>{
    let tc='tag-primary';
    if(r.result_type=='危机型')tc='tag-danger';
    else if(r.result_type=='问题型')tc='tag-warning';
    else if(r.result_type=='潜力型')tc='tag-info';
    else if(r.result_type=='健康型')tc='tag-success';
    return '<tr>'+[r.name,r.company,r.phone,r.industry,r.score_total,r.score_d1,r.score_d2,r.score_d3,r.score_d4,r.score_d5,
      '<span class="'+tc+'">'+r.result_type+'</span>',
      (r.created_at||'').replace('T',' ').slice(0,16)
    ].map(v=>'<td>'+v+'</td>').join('')+'</tr>';
  }).join('');
  document.getElementById('tbody').innerHTML=html||'<tr><td colspan="12" style="color:#999;padding:32px">暂无数据</td></tr>';
}
load();
document.getElementById('kw').addEventListener('keyup',e=>{if(e.key=='Enter')render()});
</script>
</body></html>'''


# ===== 工具函数 =====
def _parse_body(body_raw):
    """解析 request body，支持 JSON 和 form-urlencoded"""
    if isinstance(body_raw, dict):
        return body_raw
    if isinstance(body_raw, str) and body_raw:
        # 尝试 JSON
        try:
            return json.loads(body_raw)
        except Exception:
            pass
        # 尝试 form-urlencoded
        result = {}
        for pair in body_raw.split('&'):
            if '=' in pair:
                k, v = pair.split('=', 1)
                result[k] = v
        return result
    return {}


def _get_result_type(score):
    if score <= 10:
        return '危机型'
    elif score <= 20:
        return '问题型'
    elif score <= 30:
        return '潜力型'
    elif score <= 40:
        return '健康型'
    else:
        return '优秀型'


def _is_mobile(ua):
    return any(k in (ua or '').lower() for k in ['mobile', 'android', 'iphone', 'ipad', 'phone'])


# ===== 主入口 =====
def handler(request):
    method = (request.get('method') or request.get('httpMethod') or 'GET').upper()
    path = request.get('path', '/')
    headers_in = request.get('headers', {})
    cookies = headers_in.get('Cookie', '') or headers_in.get('cookie', '')
    body_raw = request.get('body', '')

    body = _parse_body(body_raw)

    # /api/health
    if path == '/api/health':
        return _json_resp({'status': 'ok', 'db': bool(AIRTABLE_API_KEY and AIRTABLE_BASE_ID)})

    # / - 首页
    if path == '/' or path == '':
        ua = headers_in.get('User-Agent', '')
        if _is_mobile(ua):
            html = _static('index_mobile.html')
        else:
            html = _static('index.html')
        if html:
            return _html_resp(html)
        return _html_resp('Not Found', 404)

    # /mobile /desktop
    if path == '/mobile':
        html = _static('index_mobile.html')
        return _html_resp(html or 'Not Found', 200 if html else 404)
    if path == '/desktop':
        html = _static('index.html')
        return _html_resp(html or 'Not Found', 200 if html else 404)

    # /admin/login
    if path == '/admin/login':
        if method == 'POST':
            pw = body.get('password', '')
            if pw == ADMIN_PASSWORD:
                return {
                    'statusCode': 200,
                    'headers': {
                        'Content-Type': 'application/json; charset=utf-8',
                        'Set-Cookie': _make_cookie({'auth': True})
                    },
                    'body': json.dumps({'success': True}, ensure_ascii=False)
                }
            return _json_resp({'success': False, 'message': '密码错误'}, 401)
        # GET
        html = LOGIN_HTML.replace('{{ERROR}}', '')
        return _html_resp(html)

    # /admin/logout
    if path == '/admin/logout':
        return {
            'statusCode': 302,
            'headers': {
                'Location': '/admin/login',
                'Set-Cookie': SESSION_COOKIE + '=; Path=/; Max-Age=0'
            },
            'body': ''
        }

    # /admin
    if path == '/admin':
        if not _is_auth(cookies):
            return _redirect_resp('/admin/login')
        return _html_resp(ADMIN_HTML)

    # /api/submit
    if path == '/api/submit':
        if method != 'POST':
            return _json_resp({'success': False, 'message': 'Method not allowed'}, 405)
        answers = body.get('answers', {})
        if not answers or len(answers) < 50:
            return _json_resp({'success': False, 'message': '请完成所有题目'}, 400)

        sd = []
        for start in range(1, 42, 10):
            cnt = 0
            for i in range(start, start + 10):
                if answers.get(str(i)) == 'yes':
                    cnt += 1
            sd.append(cnt)
        total = sum(sd)
        rtype = _get_result_type(total)

        if AIRTABLE_API_KEY and AIRTABLE_BASE_ID:
            try:
                r = req_lib.post(
                    _airtable_url(),
                    headers=_airtable_headers(),
                    json={'fields': {
                        'Name': body.get('name', ''),
                        'Company': body.get('company', ''),
                        'Phone': body.get('phone', ''),
                        'Industry': body.get('industry', ''),
                        'Answers': json.dumps(answers, ensure_ascii=False),
                        'ScoreTotal': total,
                        'ScoreD1': sd[0], 'ScoreD2': sd[1], 'ScoreD3': sd[2],
                        'ScoreD4': sd[3], 'ScoreD5': sd[4],
                        'ResultType': rtype,
                        'CreatedAt': datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%S.000Z'),
                    }},
                    timeout=10
                )
                if r.status_code not in (200, 201):
                    print('Airtable error:', r.text)
            except Exception as e:
                print('Airtable save error:', e)

        return _json_resp({
            'success': True,
            'score_total': total,
            'score_d1': sd[0], 'score_d2': sd[1], 'score_d3': sd[2],
            'score_d4': sd[3], 'score_d5': sd[4], 'result_type': rtype
        })

    # /api/records
    if path == '/api/records':
        if not _is_auth(cookies):
            return _json_resp({'success': False, 'message': '未授权'}, 401)
        if not (AIRTABLE_API_KEY and AIRTABLE_BASE_ID):
            return _json_resp({'success': True, 'total': 0, 'records': []})
        try:
            r = req_lib.get(_airtable_url(), headers=_airtable_headers(), params={'pageSize': 100}, timeout=10)
            if r.status_code != 200:
                return _json_resp({'success': False, 'message': '读取失败'}, 500)
            recs = []
            for item in r.json().get('records', []):
                f = item.get('fields', {})
                try:
                    ans = json.loads(f.get('Answers', '{}'))
                except Exception:
                    ans = {}
                recs.append({
                    'id': item.get('id', ''),
                    'name': f.get('Name', ''),
                    'company': f.get('Company', ''),
                    'phone': f.get('Phone', ''),
                    'industry': f.get('Industry', ''),
                    'answers': ans,
                    'score_total': f.get('ScoreTotal', 0),
                    'score_d1': f.get('ScoreD1', 0), 'score_d2': f.get('ScoreD2', 0),
                    'score_d3': f.get('ScoreD3', 0), 'score_d4': f.get('ScoreD4', 0),
                    'score_d5': f.get('ScoreD5', 0),
                    'result_type': f.get('ResultType', ''),
                    'created_at': f.get('CreatedAt', ''),
                })
            return _json_resp({'success': True, 'total': len(recs), 'records': recs})
        except Exception as e:
            return _json_resp({'success': False, 'message': str(e)}, 500)

    # /api/export/excel
    if path == '/api/export/excel':
        if not _is_auth(cookies):
            return _json_resp({'success': False, 'message': '未授权'}, 401)
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            return _json_resp({'success': False, 'message': '请先安装openpyxl'}, 500)

        rows = []
        if AIRTABLE_API_KEY and AIRTABLE_BASE_ID:
            try:
                r = req_lib.get(_airtable_url(), headers=_airtable_headers(), params={'pageSize': 100}, timeout=10)
                if r.status_code == 200:
                    for item in r.json().get('records', []):
                        f = item.get('fields', {})
                        rows.append([
                            item.get('id', ''),
                            f.get('Name', ''),
                            f.get('Company', ''),
                            f.get('Phone', ''),
                            f.get('Industry', ''),
                            f.get('ScoreTotal', 0),
                            f.get('ScoreD1', 0), f.get('ScoreD2', 0),
                            f.get('ScoreD3', 0), f.get('ScoreD4', 0),
                            f.get('ScoreD5', 0),
                            f.get('ResultType', ''),
                            f.get('CreatedAt', ''),
                        ])
            except Exception as e:
                print('Export error:', e)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '自检清单数据'
        hfont = Font(bold=True, color='FFFFFF', size=11)
        hfill = PatternFill(start_color='1a365d', end_color='1a365d', fill_type='solid')
        halign = Alignment(horizontal='center', vertical='center')
        bdr = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
        hdrs = ['ID', '姓名', '公司', '电话', '行业',
                '总分', '商业模式', '品牌定位', '流量获客', '转化成交', '组织系统',
                '类型', '提交时间']
        for c, h in enumerate(hdrs, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = hfont
            cell.fill = hfill
            cell.alignment = halign
            cell.border = bdr
        for ri, row in enumerate(rows, 2):
            for ci, v in enumerate(row, 1):
                cell = ws.cell(row=ri, column=ci, value=v)
                cell.border = bdr
                cell.alignment = Alignment(horizontal='center', vertical='center')
        for col in ws.columns:
            ml = 0
            first_cell = list(col)[0]
            letter = first_cell.column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > ml:
                        ml = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[letter].width = min(ml + 2, 50)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        excel_bytes = buf.read()
        fname = '自检清单数据_' + datetime.now().strftime('%Y%m%d_%H%M%S') + '.xlsx'
        return {
            'statusCode': 200,
            'headers': {
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'Content-Disposition': 'attachment; filename="' + fname + '"',
            },
            'body': base64.b64encode(excel_bytes).decode('utf-8'),
            'encoding': 'base64',
        }

    # 404
    return _html_resp('404 Not Found', 404)

