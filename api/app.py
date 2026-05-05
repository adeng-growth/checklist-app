from flask import Flask, request, jsonify, make_response, session, redirect, render_template_string, send_file, send_from_directory
import os
import json
import requests as req_lib
from datetime import datetime
from io import BytesIO

# 当前文件所在目录（即 api/ 目录）
_HERE = os.path.dirname(os.path.abspath(__file__))

app = Flask(__name__)
app.config['JSON_AS_ASCII'] = False
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'adeng-checklist-2026')

# Airtable 配置
AIRTABLE_API_KEY = os.getenv('AIRTABLE_API_KEY', '')
AIRTABLE_BASE_ID = os.getenv('AIRTABLE_BASE_ID', '')
AIRTABLE_TABLE_NAME = 'Submissions'
ADMIN_PASSWORD = os.getenv('ADMIN_PASSWORD', 'adeng2026')


def _airtable_headers():
    return {
        'Authorization': f'Bearer {AIRTABLE_API_KEY}',
        'Content-Type': 'application/json'
    }


def _airtable_url():
    return f'https://api.airtable.com/v0/{AIRTABLE_BASE_ID}/{AIRTABLE_TABLE_NAME}'


def _get_result_type(score):
    if score <= 10: return '危机型'
    elif score <= 20: return '问题型'
    elif score <= 30: return '潜力型'
    elif score <= 40: return '健康型'
    else: return '优秀型'


def _is_mobile():
    ua = request.headers.get('User-Agent', '').lower()
    return any(k in ua for k in ['mobile', 'android', 'iphone', 'ipad', 'phone'])


# ========== 页面路由 ==========
@app.route('/')
def _index():
    if _is_mobile():
        return send_from_directory(_HERE, 'index_mobile.html')
    return send_from_directory(_HERE, 'index.html')


@app.route('/mobile')
def _mobile():
    return send_from_directory(_HERE, 'index_mobile.html')


@app.route('/desktop')
def _desktop():
    return send_from_directory(_HERE, 'index.html')


@app.route('/admin')
def _admin():
    return send_from_directory(_HERE, 'admin.html')


# ========== 登录页面 ==========
_LOGIN_HTML = r'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>后台登录</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:-apple-system,BlinkMacSystemFont,"PingFang SC","Microsoft YaHei",sans-serif;background:linear-gradient(135deg,#1a365d,#2d5a87);min-height:100vh;display:flex;align-items:center;justify-content:center}
.box{background:#fff;padding:40px 36px;border-radius:20px;box-shadow:0 20px 60px rgba(0,0,0,.3);width:92%;max-width:380px}
h2{font-size:22px;color:#1a365d;margin-bottom:8px;text-align:center}
.sub{font-size:13px;color:#999;margin-bottom:28px;text-align:center}
.err{background:#ffebee;color:#c62828;padding:10px 16px;border-radius:8px;margin-bottom:16px;font-size:13px;text-align:center}
input{width:100%;padding:12px 16px;border:1.5px solid #e5e5e5;border-radius:10px;font-size:15px;outline:none;margin-bottom:20px;transition:border .2s}
input:focus{border-color:#2d5a87}
button{width:100%;padding:14px;background:linear-gradient(135deg,#1a365d,#2d5a87);color:#fff;border:none;border-radius:10px;font-size:16px;font-weight:600;cursor:pointer}
button:active{transform:scale(.97)}
.ft{font-size:11px;color:#ccc;text-align:center;margin-top:20px}
</style>
</head>
<body>
<div class="box">
<h2>&#x1F512; 后台登录</h2>
<p class="sub">中小企业增长自检清单</p>
{% if error %}<div class="err">{{ error }}</div>{% endif %}
<form method="POST"><input type="password" name="password" placeholder="请输入后台密码" required autofocus>
<button type="submit">登 录</button></form>
<p class="ft">ADENG CHECKLIST 2026</p>
</div></body></html>'''


# ========== 认证中间件 ==========
@app.before_request
def _auth_check():
    guarded = ['/api/records', '/api/export']
    if any(request.path.startswith(p) for p in guarded):
        if not session.get('admin_authed'):
            return jsonify({'success': False, 'message': '未授权'}), 401


@app.route('/admin/login', methods=['GET', 'POST'])
def _login():
    if request.method == 'POST':
        if request.form.get('password', '') == ADMIN_PASSWORD:
            session['admin_authed'] = True
            return jsonify({'success': True})
        return jsonify({'success': False, 'message': '密码错误'}), 401
    return render_template_string(_LOGIN_HTML, error='')


@app.route('/admin/logout')
def _logout():
    session.pop('admin_authed', None)
    return redirect('/admin')


# ========== 提交答案 ==========
@app.route('/api/submit', methods=['POST'])
def _submit():
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'message': '无效请求'}), 400
    answers = data.get('answers', {})
    if len(answers) < 50:
        return jsonify({'success': False, 'message': '请完成所有题目'}), 400

    def _calc(s, e):
        return sum(1 for i in range(s, e + 1) if answers.get(str(i)) == 'yes')

    sd = [_calc(i, i + 9) for i in range(1, 42, 10)]
    total = sum(sd)
    rtype = _get_result_type(total)

    # 写入 Airtable
    if AIRTABLE_API_KEY and AIRTABLE_BASE_ID:
        try:
            r = req_lib.post(_airtable_url(), headers=_airtable_headers(), json={'fields': {
                'Name': data.get('name', ''),
                'Company': data.get('company', ''),
                'Phone': data.get('phone', ''),
                'Industry': data.get('industry', ''),
                'Answers': json.dumps(answers, ensure_ascii=False),
                'ScoreTotal': total,
                'ScoreD1': sd[0], 'ScoreD2': sd[1], 'ScoreD3': sd[2], 'ScoreD4': sd[3], 'ScoreD5': sd[4],
                'ResultType': rtype,
                'CreatedAt': datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%S.000Z'),
            }}, timeout=10)
            if r.status_code not in (200, 201):
                print('AT err:', r.text)
        except Exception as e:
            print('AT save err:', e)

    return jsonify({'success': True, 'score_total': total,
                    'score_d1': sd[0], 'score_d2': sd[1], 'score_d3': sd[2],
                    'score_d4': sd[3], 'score_d5': sd[4], 'result_type': rtype})


# ========== 查询记录 ==========
@app.route('/api/records')
def _records():
    if not (AIRTABLE_API_KEY and AIRTABLE_BASE_ID):
        return jsonify({'success': True, 'total': 0, 'records': []})
    try:
        r = req_lib.get(_airtable_url(), headers=_airtable_headers(), params={'pageSize': 100}, timeout=10)
        if r.status_code != 200:
            return jsonify({'success': False, 'message': '读取失败'}), 500
        recs = []
        for item in r.json().get('records', []):
            f = item.get('fields', {})
            try:
                ans = json.loads(f.get('Answers', '{}'))
            except Exception:
                ans = {}
            recs.append({
                'id': item.get('id', ''),
                'name': f.get('Name', ''), 'company': f.get('Company', ''),
                'phone': f.get('Phone', ''), 'industry': f.get('Industry', ''),
                'answers': ans,
                'score_total': f.get('ScoreTotal', 0),
                'score_d1': f.get('ScoreD1', 0), 'score_d2': f.get('ScoreD2', 0),
                'score_d3': f.get('ScoreD3', 0), 'score_d4': f.get('ScoreD4', 0),
                'score_d5': f.get('ScoreD5', 0),
                'result_type': f.get('ResultType', ''), 'created_at': f.get('CreatedAt', ''),
            })
        return jsonify({'success': True, 'total': len(recs), 'records': recs})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ========== 导出 Excel ==========
@app.route('/api/export/excel')
def _export():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return jsonify({'success': False, 'message': 'openpyxl missing'}), 500

    rows = []
    if AIRTABLE_API_KEY and AIRTABLE_BASE_ID:
        try:
            r = req_lib.get(_airtable_url(), headers=_airtable_headers(), params={'pageSize': 100}, timeout=10)
            if r.status_code == 200:
                for item in r.json().get('records', []):
                    f = item.get('fields', {})
                    rows.append([item.get('id', ''), f.get('Name', ''), f.get('Company', ''),
                                f.get('Phone', ''), f.get('Industry', ''), f.get('ScoreTotal', 0),
                                f.get('ScoreD1', 0), f.get('ScoreD2', 0), f.get('ScoreD3', 0),
                                f.get('ScoreD4', 0), f.get('ScoreD5', 0), f.get('ResultType', ''),
                                f.get('CreatedAt', '')])
        except Exception as e:
            print('Export err:', e)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '自检清单数据'
    hfont = Font(bold=True, color='FFFFFF', size=11)
    hfill = PatternFill(start_color='1a365d', end_color='1a365d', fill_type='solid')
    halign = Alignment(horizontal='center', vertical='center')
    bdr = Border(left=Side(style='thin'), right=Side(style='thin'),
                 top=Side(style='thin'), bottom=Side(style='thin'))
    hdrs = ['ID', '姓名', '公司', '电话', '行业', '总分', '商业模式', '品牌定位', '流量获客', '转化成交', '组织系统', '类型', '提交时间']
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = halign
        cell.border = bdr
    for ri, row in enumerate(rows, 2):
        for ci, v in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = bdr
            cell.alignment = Alignment(horizontal='center', vertical='center')
    for col in ws.columns:
        ml = 0
        letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > ml:
                    ml = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[letter].width = min(ml + 2, 50)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = 'checklist_%s.xlsx' % datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)


@app.route('/api/health')
def _health():
    return jsonify({'status': 'ok', 'db': bool(AIRTABLE_API_KEY and AIRTABLE_BASE_ID)})
