#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
中小企业增长自检清单 - Flask后端（Render部署版）
提供API接口和SQLite数据库存储，支持Excel导出
"""

import sqlite3
import json
import os
from datetime import datetime
from flask import Flask, request, jsonify, send_from_directory, make_response, session, redirect, render_template_string
import io

app = Flask(__name__, static_folder='.', static_url_path='')

# 生产环境配置
app.config['JSON_AS_ASCII'] = False
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'adeng-checklist-2026')

# 数据库路径
DB_PATH = os.path.join(os.path.dirname(__file__), 'data', 'checklist.db')

# 后台密码（可从环境变量修改，默认：adeng2026）
ADMIN_PASSWORD = os.getenv('ADMIN_PASSWORD', 'adeng2026')


def init_db():
    """初始化数据库"""
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS submissions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT,
            company TEXT,
            phone TEXT,
            industry TEXT,
            answers TEXT NOT NULL,
            score_total INTEGER NOT NULL,
            score_d1 INTEGER NOT NULL,
            score_d2 INTEGER NOT NULL,
            score_d3 INTEGER NOT NULL,
            score_d4 INTEGER NOT NULL,
            score_d5 INTEGER NOT NULL,
            result_type TEXT NOT NULL,
            created_at TEXT NOT NULL
        )
    ''')
    conn.commit()
    conn.close()


def get_result_type(score):
    if score <= 10:
        return '危机型'
    elif score <= 20:
        return '问题型'
    elif score <= 30:
        return '潜力型'
    elif score <= 40:
        return '健康型'
    else:
        return '优秀型'


# ========== 后台密码验证 ==========
@app.before_request
def check_admin_auth():
    """后台路由密码验证"""
    admin_paths = ['/admin', '/api/records', '/api/export']
    if any(request.path.startswith(p) for p in admin_paths):
        if not session.get('admin_authenticated'):
            if request.path != '/admin/login':
                return redirect('/admin/login')


@app.route('/admin/login', methods=['GET', 'POST'])
def login():
    """后台登录页面"""
    if request.method == 'POST':
        password = request.form.get('password', '')
        if password == ADMIN_PASSWORD:
            session['admin_authenticated'] = True
            return redirect('/admin')
        else:
            return render_template_string(LOGIN_HTML, error='密码错误，请重试')
    
    # GET 请求，显示登录页面
    return render_template_string(LOGIN_HTML, error='')


@app.route('/admin/logout')
def logout():
    """退出登录"""
    session.pop('admin_authenticated', None)
    return redirect('/admin/login')


# 登录页面HTML
LOGIN_HTML = '''
<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>后台登录 - 中小企业自检清单</title>
<style>
*{margin:0;padding:0;box-sizing:border-box;}
body{
  font-family:-apple-system,BlinkMacSystemFont,"PingFang SC","Microsoft YaHei",sans-serif;
  background:linear-gradient(135deg,#1a365d 0%,#2d5a87 100%);
  min-height:100vh;
  display:flex;align-items:center;justify-content:center;
}
.login-box{
  background:#fff;
  padding:40px 36px;
  border-radius:20px;
  box-shadow:0 20px 60px rgba(0,0,0,.3);
  width:92%;max-width:380px;
}
h2{font-size:22px;color:#1a365d;margin-bottom:8px;text-align:center;}
.sub{font-size:13px;color:#999;margin-bottom:28px;text-align:center;}
.error{background:#ffebee;color:#c62828;padding:10px 16px;border-radius:8px;margin-bottom:16px;font-size:13px;text-align:center;}
input{
  width:100%;padding:12px 16px;
  border:1.5px solid #e5e5e5;border-radius:10px;
  font-size:15px;outline:none;margin-bottom:20px;
  transition:border-color .2s;
}
input:focus{border-color:#2d5a87;}
button{
  width:100%;padding:14px;
  background:linear-gradient(135deg,#1a365d,#2d5a87);
  color:#fff;border:none;border-radius:10px;
  font-size:16px;font-weight:600;cursor:pointer;
  transition:transform .15s;
}
button:active{transform:scale(.97);}
.footer{font-size:11px;color:#ccc;text-align:center;margin-top:20px;}
</style>
</head>
<body>
<div class="login-box">
  <h2>🔒 后台登录</h2>
  <p class="sub">中小企业增长自检清单</p>
  {% if error %}<div class="error">{{ error }}</div>{% endif %}
  <form method="POST">
    <input type="password" name="password" placeholder="请输入后台密码" required autofocus>
    <button type="submit">登 录</button>
  </form>
  <p class="footer">ADENG CHECKLIST · 2026</p>
</div>
</body>
</html>
'''


# ========== 前台页面 ==========
def check_mobile():
    """检查是否为移动设备"""
    user_agent = request.headers.get('User-Agent', '').lower()
    keywords = ['mobile', 'android', 'iphone', 'ipad', 'phone']
    return any(k in user_agent for k in keywords)

@app.route('/')
def index():
    """自动判断设备类型，返回对应页面"""
    if check_mobile():
        return send_from_directory('.', 'index_mobile.html')
    else:
        return send_from_directory('.', 'index.html')

@app.route('/mobile')
def mobile():
    """强制移动版"""
    return send_from_directory('.', 'index_mobile.html')

@app.route('/desktop')
def desktop():
    """强制桌面版"""
    return send_from_directory('.', 'index.html')


# ========== API接口 ==========
@app.route('/api/submit', methods=['POST'])
def submit():
    """提交答题结果"""
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'message': '无效请求'}), 400

    answers = data.get('answers', {})
    if len(answers) < 50:
        return jsonify({'success': False, 'message': '请完成所有题目'}), 400

    # 计算各维度得分
    def calc_score(start, end):
        return sum(1 for i in range(start, end + 1) if answers.get(str(i)) == 'yes')

    score_d1 = calc_score(1, 10)
    score_d2 = calc_score(11, 20)
    score_d3 = calc_score(21, 30)
    score_d4 = calc_score(31, 40)
    score_d5 = calc_score(41, 50)
    score_total = score_d1 + score_d2 + score_d3 + score_d4 + score_d5
    result_type = get_result_type(score_total)

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''
        INSERT INTO submissions 
        (name, company, phone, industry, answers, score_total, score_d1, score_d2, score_d3, score_d4, score_d5, result_type, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        data.get('name', ''),
        data.get('company', ''),
        data.get('phone', ''),
        data.get('industry', ''),
        json.dumps(answers, ensure_ascii=False),
        score_total,
        score_d1, score_d2, score_d3, score_d4, score_d5,
        result_type,
        datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ))
    conn.commit()
    submission_id = c.lastrowid
    conn.close()

    return jsonify({
        'success': True,
        'id': submission_id,
        'score_total': score_total,
        'score_d1': score_d1,
        'score_d2': score_d2,
        'score_d3': score_d3,
        'score_d4': score_d4,
        'score_d5': score_d5,
        'result_type': result_type
    })


@app.route('/api/records', methods=['GET'])
def get_records():
    """获取所有提交记录（后台管理）"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('SELECT * FROM submissions ORDER BY created_at DESC')
    rows = c.fetchall()
    conn.close()

    columns = ['id', 'name', 'company', 'phone', 'industry', 'answers',
               'score_total', 'score_d1', 'score_d2', 'score_d3', 'score_d4', 'score_d5',
               'result_type', 'created_at']
    records = [dict(zip(columns, row)) for row in rows]

    # 解析 answers JSON
    for r in records:
        if r['answers']:
            r['answers'] = json.loads(r['answers'])

    return jsonify({
        'success': True,
        'total': len(records),
        'records': records
    })


@app.route('/api/export/excel')
def export_excel():
    """导出Excel文件"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return jsonify({'success': False, 'message': '请先安装 openpyxl: pip install openpyxl'}), 500

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('SELECT * FROM submissions ORDER BY created_at DESC')
    rows = c.fetchall()
    conn.close()

    # 创建Excel工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "自检清单数据"

    # 标题行样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 写入标题行
    headers = ['ID', '姓名', '公司', '电话', '行业', 
               '总分', '商业模式', '品牌定位', '流量获客', '转化成交', '组织系统',
               '类型', '提交时间']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 写入数据行
    for row_idx, row in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=row[0])
        ws.cell(row=row_idx, column=2, value=row[1])
        ws.cell(row=row_idx, column=3, value=row[2])
        ws.cell(row=row_idx, column=4, value=row[3])
        ws.cell(row=row_idx, column=5, value=row[4])
        ws.cell(row=row_idx, column=6, value=row[6])
        ws.cell(row=row_idx, column=7, value=row[7])
        ws.cell(row=row_idx, column=8, value=row[8])
        ws.cell(row=row_idx, column=9, value=row[9])
        ws.cell(row=row_idx, column=10, value=row[10])
        ws.cell(row=row_idx, column=11, value=row[11])
        ws.cell(row=row_idx, column=12, value=row[12])
        ws.cell(row=row_idx, column=13, value=row[13])

        # 设置边框
        for col in range(1, 14):
            ws.cell(row=row_idx, column=col).border = thin_border
            ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal="center", vertical="center")

    # 自动调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # 保存到内存
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # 返回文件
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"自检清单数据_{timestamp}.xlsx"
    
    response = make_response(send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    ))
    
    return response


# ========== 后台管理页面 ==========
@app.route('/admin')
def admin():
    """后台管理页面"""
    if not session.get('admin_authenticated'):
        return redirect('/admin/login')
    return send_from_directory('.', 'admin.html')


if __name__ == '__main__':
    init_db()
    print('=' * 50)
    print('服务器已启动！')
    print('前台地址: http://localhost:5050')
    print('后台地址: http://localhost:5050/admin')
    print('后台密码: adeng2026')
    print('=' * 50)
    app.run(host='0.0.0.0', port=5050, debug=False)
